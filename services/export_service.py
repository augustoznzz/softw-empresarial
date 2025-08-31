#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Serviço de exportação para PDF e Excel
"""

import os
from datetime import datetime
from typing import List, Dict, Any
from models.imovel import Imovel
from models.database import DatabaseManager
from utils.formatacao import formatar_moeda
import logging

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab não disponível - exportação PDF desabilitada")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logging.warning("OpenPyXL não disponível - exportação Excel desabilitada")

class ExportService:
    def __init__(self):
        self.db_manager = DatabaseManager()
        
    def export_to_pdf(self, imoveis: List[Imovel], filepath: str, filtros: Dict[str, Any] = None) -> bool:
        """Exporta dados para PDF"""
        if not REPORTLAB_AVAILABLE:
            logging.error("ReportLab não disponível para exportação PDF")
            return False
            
        try:
            # Criar documento PDF
            doc = SimpleDocTemplate(filepath, pagesize=A4)
            story = []
            
            # Estilos
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER
            )
            
            # Título
            title = Paragraph("Relatório de Imóveis - Sistema de Negociação", title_style)
            story.append(title)
            
            # Data e hora
            data_hora = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", styles['Normal'])
            story.append(data_hora)
            story.append(Spacer(1, 20))
            
            # Filtros aplicados
            if filtros:
                filtros_text = "Filtros aplicados: "
                filtros_list = []
                for key, value in filtros.items():
                    if value:
                        filtros_list.append(f"{key}: {value}")
                if filtros_list:
                    filtros_text += ", ".join(filtros_list)
                    filtros_para = Paragraph(filtros_text, styles['Normal'])
                    story.append(filtros_para)
                    story.append(Spacer(1, 20))
            
            # Resumo
            total_imoveis = len(imoveis)
            story.append(Paragraph(f"Total de imóveis: {total_imoveis}", styles['Heading2']))
            story.append(Spacer(1, 20))
            
            # Tabela de imóveis
            if imoveis:
                # Cabeçalhos - Removido endereço, mantido apenas CEP
                headers = [
                    'CEP', 'Cidade', 'Metragem', 'Custo Total', 
                    'Preço Estimado', 'Margem', 'ROI (%)', 'Status'
                ]
                
                # Dados
                data = [headers]
                for imovel in imoveis:
                    # Calcular valores
                    custo_total = imovel.get_custo_total()
                    preco_estimado = self._calcular_preco_estimado(imovel)
                    margem = preco_estimado - custo_total
                    roi = (margem / custo_total * 100) if custo_total > 0 else 0
                    
                    row = [
                        imovel.cep,  # Removido endereço, mantido apenas CEP
                        imovel.cidade,
                        f"{imovel.metragem:.1f} m²",
                        formatar_moeda(custo_total),
                        formatar_moeda(preco_estimado),
                        formatar_moeda(margem),
                        f"{roi:.1f}%",
                        imovel.status.replace('_', ' ').title()
                    ]
                    data.append(row)
                
                # Criar tabela
                table = Table(data, colWidths=[1.5*inch, 1*inch, 0.8*inch, 1.2*inch, 1.2*inch, 1*inch, 0.6*inch, 0.8*inch])
                
                # Estilo da tabela
                table_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ])
                
                table.setStyle(table_style)
                story.append(table)
                
                # Resumo financeiro
                story.append(Spacer(1, 30))
                story.append(Paragraph("Resumo Financeiro", styles['Heading2']))
                
                total_custo = sum(imovel.get_custo_total() for imovel in imoveis)
                total_preco_estimado = sum(self._calcular_preco_estimado(imovel) for imovel in imoveis)
                total_margem = total_preco_estimado - total_custo
                roi_medio = (total_margem / total_custo * 100) if total_custo > 0 else 0
                
                resumo_data = [
                    ['Total Custo', 'Total Preço\nEstimado', 'Total Margem', 'ROI Médio'],
                    [
                        formatar_moeda(total_custo),
                        formatar_moeda(total_preco_estimado),
                        formatar_moeda(total_margem),
                        f"{roi_medio:.1f}%"
                    ]
                ]
                
                # Ajustar larguras das colunas: expandir "Total Preço Estimado" e reduzir 5% "ROI Médio"
                resumo_table = Table(resumo_data, colWidths=[1.4*inch, 2.0*inch, 1.4*inch, 1.425*inch])
                resumo_style = TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 11),  # Reduzir fonte do cabeçalho para 11
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('WORDWRAP', (0, 0), (-1, -1), True),  # Permitir quebra de linha
                ])
                
                resumo_table.setStyle(resumo_style)
                story.append(resumo_table)
            
            # Gerar PDF
            doc.build(story)
            return True
            
        except Exception as e:
            logging.error(f"Erro ao exportar para PDF: {e}")
            return False
            
    def export_to_excel(self, imoveis: List[Imovel], filepath: str, filtros: Dict[str, Any] = None) -> bool:
        """Exporta dados para Excel"""
        if not OPENPYXL_AVAILABLE:
            logging.error("OpenPyXL não disponível para exportação Excel")
            return False
            
        try:
            # Criar workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Imóveis"
            
            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Cabeçalhos - Removido ID e endereço, reorganizada ordem
            headers = [
                'CEP', 'Cidade', 'Estado', 'Metragem', 'Quartos', 'Banheiros',
                'Ano', 'Padrão', 'Custo Aquisição', 'Custos Reforma', 'Custos Transação',
                'Custo Total', 'Preço Estimado', 'Margem', 'ROI (%)', 'Status'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                
            # Dados
            for row, imovel in enumerate(imoveis, 2):
                # Calcular valores
                custo_total = imovel.get_custo_total()
                preco_estimado = self._calcular_preco_estimado(imovel)
                margem = preco_estimado - custo_total
                roi = (margem / custo_total * 100) if custo_total > 0 else 0
                
                # Preencher linha - Removido ID e endereço, reorganizadas colunas
                ws.cell(row=row, column=1, value=imovel.cep)  # CEP agora na coluna 1
                ws.cell(row=row, column=2, value=imovel.cidade)  # Cidade agora na coluna 2
                ws.cell(row=row, column=3, value=imovel.estado)  # Estado agora na coluna 3
                ws.cell(row=row, column=4, value=imovel.metragem)  # Metragem agora na coluna 4
                ws.cell(row=row, column=5, value=imovel.quartos)  # Quartos agora na coluna 5
                ws.cell(row=row, column=6, value=imovel.banheiros)  # Banheiros agora na coluna 6
                ws.cell(row=row, column=7, value=imovel.ano)  # Ano agora na coluna 7
                ws.cell(row=row, column=8, value=imovel.padrao_acabamento)  # Padrão agora na coluna 8
                ws.cell(row=row, column=9, value=imovel.custo_aquisicao)  # Custo Aquisição agora na coluna 9
                ws.cell(row=row, column=10, value=imovel.custos_reforma)  # Custos Reforma agora na coluna 10
                ws.cell(row=row, column=11, value=imovel.custos_transacao)  # Custos Transação agora na coluna 11
                ws.cell(row=row, column=12, value=custo_total)  # Custo Total agora na coluna 12
                ws.cell(row=row, column=13, value=preco_estimado)  # Preço Estimado agora na coluna 13
                ws.cell(row=row, column=14, value=margem)  # Margem agora na coluna 14
                ws.cell(row=row, column=15, value=roi)  # ROI agora na coluna 15
                ws.cell(row=row, column=16, value=imovel.status)  # Status agora na coluna 16
                
                # Formatar valores monetários - Índices corrigidos após remoção do ID e endereço
                for col in [9, 10, 11, 12, 13, 14]:
                    cell = ws.cell(row=row, column=col)
                    cell.number_format = 'R$ #.##0'  # Formato brasileiro: pontos como separador de milhares, sem centavos
                    
                # Formatar ROI - Corrigido índice
                ws.cell(row=row, column=15).number_format = '0.0%'
                
                # Formatar metragem - Corrigido índice
                ws.cell(row=row, column=4).number_format = '0.0'
                
            # Ajustar largura das colunas
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Adicionar filtros
            if filtros:
                ws2 = wb.create_sheet("Filtros")
                ws2.cell(row=1, column=1, value="Filtros Aplicados")
                ws2.cell(row=1, column=1).font = Font(bold=True)
                
                row = 3
                for key, value in filtros.items():
                    if value:
                        ws2.cell(row=row, column=1, value=key)
                        ws2.cell(row=row, column=2, value=str(value))
                        row += 1
                        
            # Salvar arquivo
            wb.save(filepath)
            return True
            
        except Exception as e:
            logging.error(f"Erro ao exportar para Excel: {e}")
            return False
            
    def _calcular_preco_estimado(self, imovel: Imovel) -> float:
        """Calcula preço estimado para exportação (versão simplificada)"""
        try:
            # Fator de localização padrão
            fator_localizacao = 1.0
            
            # Fator de padrão
            fator_padrao = {
                'baixo': 0.9,
                'medio': 1.0,
                'alto': 1.1
            }.get(imovel.padrao_acabamento, 1.0)
            
            # Preço base (valor padrão)
            preco_base_m2 = 5000.0
            
            preco_estimado = preco_base_m2 * imovel.metragem * fator_localizacao * fator_padrao
            return round(preco_estimado, 2)
            
        except Exception as e:
            logging.error(f"Erro ao calcular preço estimado: {e}")
            return 0.0
            
    def get_export_formats(self) -> List[str]:
        """Retorna formatos de exportação disponíveis"""
        formats = []
        
        if REPORTLAB_AVAILABLE:
            formats.append("PDF")
            
        if OPENPYXL_AVAILABLE:
            formats.append("Excel")
            
        return formats
        
    def get_available_formats(self) -> List[str]:
        """Alias para get_export_formats"""
        return self.get_export_formats()
